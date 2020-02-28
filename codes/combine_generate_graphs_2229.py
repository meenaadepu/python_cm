import pandas as pd
import os
from openpyxl import load_workbook

outpath = r"D:\Meena\Ltv\22LTV\2228_Die1\outputs\rosc_delay\Final_Graphs"
out_fname = "graph_out.xlsx"

dir = r"D:\Meena\Ltv\22LTV\2228_Die1\outputs\rosc_delay\Graphs"

#cells =['CSC32', 'CSC20', 'CSC24', 'CSC36', 'CSC28']
cells =['CSC20', 'CSC24','CSC28']
df_out = pd.DataFrame()


def get_dirs(dir):
    l = []
    for root, dirs, _ in os.walk(dir):
        for d in dirs:
            s_dir = os.path.join(root, d)
            l.append(s_dir)
    print(l)
    return l[2:]

dirs = get_dirs(dir)
print(dirs)

for d in dirs:
    folder = d.split("Graphs\\")[1].split("\\")[0]
    f1name = d.split("\\")[-1].replace(" ", "_")
    out_dir =os.path.join(outpath,folder, f1name)

    if not os.path.exists(out_dir):
        os.makedirs(out_dir)
    outpath1 = os.path.join(out_dir,out_fname)

    writer = pd.ExcelWriter(outpath1, 'xlsxwriter')
    print(d)
    for i in cells:
        for path, subdirs, files in os.walk(d):
            print("files",files,)
            for name in files:
                fname = (os.path.join(path, name))
                print("fname", fname)
                j = i + d.split("_")[-1].split("VT")[0]
                print("fname",fname,j)
                wb = load_workbook(fname, read_only=True)
                print(wb.sheetnames)
                if j in wb.sheetnames:
                    df = pd.read_excel(fname, sheet_name=j)
                    df_out = pd.concat([df_out,df],axis = 1)
                else:
                    print(j,df_out.shape)

            if (len(files)>1)  :
                df_out = df_out.drop_duplicates(subset='Logic_Cell_used')
                df_op = df_out.iloc[:,1:]
                print(df_op.head(3))
                df_op = df_op.drop(['Logic_Cell_used'], axis=1)
            if df_out.empty:
                df_out = pd.DataFrame(columns=['Logic_Cell_used'])
                df_f['Logic_Cell_used'] = df_out['Logic_Cell_used']
                df_op = df_out.iloc[:, 1:]
                print(df_op.head(2))
                #df_op = df_f

            #df_out = pd.read_excel(fname)
            df_f = df_out

            df_f = pd.concat([df_out["Logic_Cell_used"],df_op],axis = 1)
                # print(list(df_f))
            df_f = df_f.loc[:, ~df_f.columns.duplicated()] #removing Logic_Cell_used multiple(duplicated) columns with single column

                # df_f = df_f.iloc[:, 2:]

            sht_name = "%s" % (j)
            #print("sht_name   ",sht_name)

            df_f.to_excel(writer, sheet_name=sht_name, index=False)

            # Access the XlsxWriter workbook and worksheet objects from the dataframe.
            workbook = writer.book
            worksheet = writer.sheets[sht_name]

            percent_fmt = workbook.add_format({'num_format': '0.00%'})
            worksheet.set_column('B:F', None, percent_fmt)

            # Create a chart object.
            chart = workbook.add_chart({'type': 'line', 'color': 'red'})

            mark_color = {'FF_VDD0.88V_VNW0V_VPW0V_125C': ['circle', '#f032e6'],
                          'FF_VDD0.88V_VNW0.8V_VPWN0.8V_125C': ['circle', '#ADFF2F'],
                          'FF_VDD0.88V_VNW0.45V_VPWN0.45V_125C': ['circle', '#00BFFF'],
                          'FF_VDD0.88V_VNW0V_VPW0V_N40C': ['circle', '#F08080'],
                          'FF_VDD0.88V_VNW0.8V_VPWN0.8V_N40C': ['circle', '#FF4500'],
                          'FF_VDD0.88V_VNW0.45V_VPWN0.45V_N40C': ['circle', '#008080'],
                          'FF_VDD0.945V_VNW0V_VPW0V_N40C': ['circle', '#00cccc'],
                          'FF_VDD0.945V_VNW0V_VPW0V_125C': ['circle', '#669900'],
                          'FF_VDD0.72V_VNW0.05V_VPWN0.05V_N40C': ['circle', '#00cccc'],
                          'FF_VDD0.72V_VNW0.05V_VPWN0.05V_125C': ['circle', '#669900'],
                          'FF_VDD0.945V_VNW0.9V_VPWN0.9V_N40C': ['circle', '#f032e6'],
                          'FF_VDD0.945V_VNW0.9V_VPWN0.9V_125C': ['circle', '#ADFF2F'],

                          'TT_VDD0.8V_VNW0.45V_VPWN0.45V_25C': ['diamond', '#4B0082'],
                          'TT_VDD0.8V_VNW0.45V_VPWN0.45V_85C': ['diamond', '#C71585'],
                          'TT_VDD0.8V_VNW0V_VPW0V_25C': ['diamond', '#9400D3'],
                          'TT_VDD0.8V_VNW0V_VPW0V_85C': ['diamond', '#0000FF'],
                          'TT_VDD0.8V_VNW0.8V_VPWN0.8V_25C': ['diamond', '#8B0000'],
                          'TT_VDD0.8V_VNW0.8V_VPWN0.8V_85C': ['diamond', '#FF8C00'],
                          'TT_VDD0.72V_VNW0.45V_VPWN0.45V_125C': ['diamond', '#B9C795'],
                          'TT_VDD0.72V_VNW0.45V_VPWN0.45V_N40C': ['diamond', '#00ff80'],
                          'TT_VDD0.72V_VNW0V_VPW0V_125C': ['diamond', '#e6194B'],
                          'TT_VDD0.72V_VNW0V_VPW0V_N40C': ['diamond', '#483D8B'],
                          'TT_VDD0.9V_VNW0V_VPW0V_25C': ['diamond', '#9400D3'],
                          'TT_VDD0.9V_VNW0V_VPW0V_85C': ['diamond', '#0000FF'],

                          'SS_VDD0.72V_VNW0.45V_VPWN0.45V_125C': ['triangle', '#a9a9a9'],
                          'SS_VDD0.72V_VNW0.45V_VPWN0.45V_N40C': ['triangle', '#f53b22'],
                          'SS_VDD0.72V_VNW0.8V_VPWN0.8V_N40C': ['triangle', '#22dcf5'],
                          'SS_VDD0.72V_VNW0.8V_VPWN0.8V_125C': ['triangle', '#f5228f'],
                          'SS_VDD0.72V_VNW0V_VPW0V_125C': ['triangle', '#cc9900'],
                          'SS_VDD0.72V_VNW0V_VPW0V_N40C': ['triangle', '#20B2AA'],
                          'SS_VDD0.76V_VNW0.45V_VPWN0.45V_N40C': ['triangle', '#000075'],
                          'SS_VDD0.76V_VNW0.45V_VPWN0.45V_125C': ['triangle', '#FFD700'],
                          'SS_VDD0.76V_VNW0.8V_VPWN0.8V_N40C': ['triangle', '#DC143C'],
                          'SS_VDD0.76V_VNW0.8V_VPWN0.8V_125C': ['triangle', '#EE82EE'],
                          'SS_VDD0.76V_VNW0V_VPW0V_N40C': ['triangle', '#000000'],
                          'SS_VDD0.76V_VNW0V_VPW0V_125C': ['triangle', '#95B9C7'],
                          'SS_VDD0.81V_VNW0V_VPW0V_N40C': ['triangle', '#22dcf5'],
                          'SS_VDD0.81V_VNW0V_VPW0V_125C': ['triangle', '#f5228f'],
                          'SS_VDD0.9V_VNW0.45V_VPWN0.45V_125C': ['triangle', '#a9a9a9'],
                          'SS_VDD0.9V_VNW0.45V_VPWN0.45V_N40C': ['triangle', '#f53b22'],

                          'FF_VDD0.72V_VNW0V_VPW0V_N40C': ['circle', '#f032e6'],
                          'FF_VDD0.72V_VNW0V_VPW0V_125C': ['circle', '#ADFF2F'],
                          'FF_VDD0.72V_VNW0.8V_VPWN0.8V_N40C': ['circle', '#00BFFF'],
                          'FF_VDD0.72V_VNW0.8V_VPWN0.8V_125C': ['circle', '#f08dcc'],
                          'FF_VDD0.72V_VNW0.45V_VPWN0.45V_N40C': ['circle', '#FF4500'],
                          'FF_VDD0.72V_VNW0.45V_VPWN0.45V_125C': ['circle', '#008080'],
                          'FF_VDD0.945V_VNW0.5V_VPWN0.5V_N40C': ['circle', '#FF4500'],
                          'FF_VDD0.945V_VNW0.5V_VPWN0.5V_125C': ['circle', '#008080'],

                          'TT_VDD0.65V_VNW0.8V_VPWN0.8V_25C': ['diamond', '#4B0082'],
                          'TT_VDD0.65V_VNW0.8V_VPWN0.8V_85C': ['diamond', '#C71585'],
                          'TT_VDD0.65V_VNW0.45V_VPWN0.45V_25C': ['diamond', '#9400D3'],
                          'TT_VDD0.65V_VNW0.45V_VPWN0.45V_85C': ['diamond', '#0000FF'],
                          'TT_VDD0.65V_VNW0V_VPW0V_25C': ['diamond', '#8B0000'],
                          'TT_VDD0.65V_VNW0V_VPW0V_85C': ['diamond', '#FF8C00'],
                          'TT_VDD0.65V_VNW0.2V_VPWN0.2V_25C': ['diamond', '#B9C795'],
                          'TT_VDD0.65V_VNW0.2V_VPWN0.2V_85C': ['diamond', '#00ff80'],
                          'TT_VDD0.59V_VNW0V_VPW0V_N40C': ['diamond', '#e6194B'],
                          'TT_VDD0.59V_VNW0V_VPW0V_125C': ['diamond', '#483D8B'],
                          'TT_VDD0.9V_VNW0.5V_VPWN0.5V_25C': ['diamond', '#483D8B'],
                          'TT_VDD0.9V_VNW0.5V_VPWN0.5V_85C': ['diamond', '#483D8B'],
                          'SS_VDD0.81V_VNW0.5V_VPWN0.5V_N40C': ['triangle', '#a9a9a9'],
                          'SS_VDD0.81V_VNW0.5V_VPWN0.5V_125C': ['triangle', '#f53b22'],
                          'TT_VDD0.9V_VNW0.9V_VPWN0.9V_25C': ['diamond', '#FFA07A'],
                          'TT_VDD0.9V_VNW0.9V_VPWN0.9V_85C': ['diamond', '#483D8B'],

                          'SS_VDD0.59V_VNW0.45V_VPWN0.45V_N40C': ['triangle', '#a9a9a9'],
                          'SS_VDD0.59V_VNW0.45V_VPWN0.45V_125C': ['triangle', '#f53b22'],
                          'SS_VDD0.59V_VNW0.8V_VPWN0.8V_125C': ['triangle', '#22dcf5'],
                          'SS_VDD0.59V_VNW0.8V_VPWN0.8V_N40C': ['triangle', '#FFD700'],
                          'SS_VDD0.59V_VNW0V_VPW0V_N40C': ['triangle', '#1d3b47'],
                          'SS_VDD0.59V_VNW0V_VPW0V_125C': ['triangle', '#EE82EE'],
                          'SS_VDD0.59V_VNW0.85V_VPWN1.15V_N40C': ['triangle', '#95B9C7'],
                          'SS_VDD0.59V_VNW0.85V_VPWN1.15V_125C': ['triangle', '#000075'],
                          'SS_VDD0.81V_VNW0.9V_VPWN0.9V_125C': ['triangle', '#B22222'],
                          'SS_VDD0.81V_VNW0.9V_VPWN0.9V_N40C': ['triangle', '#FF7F50'],

                          }


            for i in range(len(list(df_f)) - 1):
               # print(df_f.head(4))
                row = df_f.shape[0]
                col = i + 1
                s = list(df_f)[col]
                m_c_key = str(s)
                marker, color = mark_color.get(m_c_key)
                chart.add_series({
                    'name': [sht_name, 0, col],
                    'categories': [sht_name, 1, 0, row, 0],
                    'values': [sht_name, 1, col, row, col],
                    'marker': {'type': marker, 'size': 7, 'line': {'color': color}, 'border': {'color': color},
                               'fill': {'color': color}},
                    'line': {'none': True}
                })
            # Configure the chart dimensions.
            chart.set_size({'width': 1400.00, 'height': 770.00, })
            chart.set_plotarea({'layout': {'x': 0.1, 'y': 0.22, 'width': 0.99, 'height': 0.5, }})

            # Configure the chart title.
            title = "%s CL%s ROSC MHC variance (%s)" % (folder.replace("_"," "),j[3:5],f1name.replace("_"," "))

            chart.set_title({'name': title,
                             'name_font': {'name': 'Arial', 'color': 'black', 'size': 14, 'bold': True},
                             'overlay': True,
                             })

            # Configure the chart axes.
            chart.set_x_axis({'num_font': {'rotation': -45, 'name': 'Arial', 'size': 8, 'bold': True},
                              'major_gridlines': {'visible': True},
                              'label_position': 'low',
                              'name': 'ROSC Cell Name',
                              'name_font': {'name': 'Arial', 'size': 10, 'bold': True},
                              })

            chart.set_y_axis({'major_gridlines': {'visible': False},
                              'num_font': {'name': 'Arial', 'size': 10, 'bold': True},
                              'name': 'MHC Deviation in %',
                              'name_font': {'name': 'Arial', 'size': 10, 'bold': True},
                              })

            # Position the legend at the top of the chart.
            chart.set_legend({'layout': {'x': 0.03, 'y': 0.10, 'width': 0.95, 'height': 0.10, },
                              'font': {'name': 'Arial', 'size': 10}})

            # Insert the chart into the worksheet.
            worksheet.insert_chart('I2', chart)

        df_out =df_f=df_op=pd.DataFrame()
        print('Finished Processing')

    writer.save()
    writer.close()




